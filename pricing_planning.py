#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
قسم تخطيط التسعير للحملات - النسخة المحسنة
Pricing Planning Module for Campaigns - Improved Version
"""

import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime
from openai import OpenAI
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================
# 💰 تخطيط التسعير للحملات
# ============================================
def pricing_planning():
    """صفحة تخطيط التسعير للحملات"""
    
    st.markdown("""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 1.5rem; border-radius: 15px; color: white; text-align: center; margin-bottom: 2rem;'>
            <h1 style='margin: 0; font-size: 2rem;'>💰 تخطيط التسعير للحملات</h1>
            <p style='margin: 0.5rem 0 0 0; opacity: 0.9;'>خطط أسعارك بذكاء واحصل على نصائح من الذكاء الاصطناعي</p>
        </div>
    """, unsafe_allow_html=True)
    
    # تحميل البيانات
    pricing_plans = load_pricing_plans()
    
    # التبويبات
    tab1, tab2 = st.tabs(["📋 خطط التسعير", "➕ خطة جديدة"])
    
    with tab1:
        show_pricing_plans_list(pricing_plans)
    
    with tab2:
        create_new_pricing_plan()


def load_pricing_plans():
    """تحميل خطط التسعير من الملف"""
    try:
        if os.path.exists("pricing_plans.json"):
            with open("pricing_plans.json", "r", encoding="utf-8") as f:
                return json.load(f)
        return {"plans": []}
    except Exception as e:
        st.error(f"خطأ في تحميل خطط التسعير: {str(e)}")
        return {"plans": []}


def save_pricing_plans(data):
    """حفظ خطط التسعير إلى الملف"""
    try:
        with open("pricing_plans.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        st.error(f"خطأ في حفظ خطط التسعير: {str(e)}")
        return False


def show_pricing_plans_list(pricing_plans):
    """عرض قائمة خطط التسعير"""
    
    if not pricing_plans["plans"]:
        st.info("📭 لا توجد خطط تسعير بعد. ابدأ بإنشاء خطة جديدة!")
        return
    
    st.markdown(f"### 📊 إجمالي الخطط: {len(pricing_plans['plans'])}")
    
    # عرض الخطط في جدول
    for idx, plan in enumerate(pricing_plans["plans"]):
        with st.expander(f"📦 {plan['name']} - {plan.get('created_at', 'غير محدد')}"):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("📦 عدد المنتجات", plan['analytics']['total_products'])
            
            with col2:
                st.metric("💰 الإيرادات المتوقعة", f"{plan['analytics']['total_revenue']:,.0f} ر.س")
            
            with col3:
                st.metric("📈 الربح المتوقع", f"{plan['analytics']['total_profit']:,.0f} ر.س")
            
            with col4:
                st.metric("📊 متوسط الربح", f"{plan['analytics']['avg_profit_margin']:.1f}%")
            
            # أزرار الإجراءات
            col_btn1, col_btn2, col_btn3 = st.columns(3)
            
            with col_btn1:
                if st.button("👁️ عرض التفاصيل", key=f"view_{idx}", use_container_width=True):
                    show_plan_details(plan)
            
            with col_btn2:
                if st.button("📥 تصدير Excel", key=f"export_{idx}", use_container_width=True):
                    excel_data = export_plan_to_excel(plan)
                    st.download_button(
                        label="⬇️ تحميل الملف",
                        data=excel_data,
                        file_name=f"{plan['name']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{idx}"
                    )
            
            with col_btn3:
                if st.button("🗑️ حذف", key=f"delete_{idx}", use_container_width=True, type="secondary"):
                    pricing_plans["plans"].pop(idx)
                    save_pricing_plans(pricing_plans)
                    st.success("✅ تم حذف الخطة بنجاح!")
                    st.rerun()


def show_plan_details(plan):
    """عرض تفاصيل خطة التسعير"""
    st.markdown("---")
    st.markdown("### 📋 تفاصيل الخطة")
    
    # معلومات الخطة
    st.markdown(f"**الاسم:** {plan['name']}")
    st.markdown(f"**الوصف:** {plan.get('description', 'لا يوجد')}")
    st.markdown(f"**تاريخ الإنشاء:** {plan.get('created_at', 'غير محدد')}")
    st.markdown(f"**المنشئ:** {plan.get('created_by', 'غير محدد')}")
    
    # جدول المنتجات
    st.markdown("### 📦 المنتجات")
    df = pd.DataFrame(plan['products'])
    
    # تنسيق الجدول
    def color_profit_margin(val):
        if val >= 30:
            return 'background-color: #d4edda'
        elif val >= 15:
            return 'background-color: #fff3cd'
        else:
            return 'background-color: #f8d7da'
    
    styled_df = df.style.applymap(color_profit_margin, subset=['profit_margin'])
    st.dataframe(styled_df, use_container_width=True, height=400)


def create_new_pricing_plan():
    """إنشاء خطة تسعير جديدة"""
    
    st.markdown("### ➕ إنشاء خطة تسعير جديدة")
    
    # المرحلة 1: معلومات الخطة
    st.markdown("#### 1️⃣ معلومات الخطة")
    
    col1, col2 = st.columns(2)
    
    with col1:
        plan_name = st.text_input(
            "📝 اسم الخطة:",
            placeholder="مثال: خطة رمضان 2024",
            key="plan_name"
        )
    
    with col2:
        # تحميل الحملات
        campaigns = load_campaigns()
        campaign_options = ["بدون ربط بحملة"] + [c['name'] for c in campaigns.get('campaigns', [])]
        selected_campaign = st.selectbox(
            "📅 الحملة المرتبطة:",
            campaign_options,
            key="selected_campaign"
        )
    
    plan_description = st.text_area(
        "📄 الوصف:",
        placeholder="اكتب وصفاً مختصراً للخطة...",
        key="plan_description"
    )
    
    # المرحلة 2: إعدادات الخصم
    st.markdown("#### 2️⃣ إعدادات الخصم العامة")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        base_discount = st.number_input(
            "💸 نسبة الخصم الأساسية (%):",
            min_value=0.0,
            max_value=100.0,
            value=30.0,
            step=1.0,
            key="base_discount"
        )
    
    with col2:
        code_discount = st.number_input(
            "🎟️ نسبة خصم الكود (%):",
            min_value=0.0,
            max_value=100.0,
            value=10.0,
            step=1.0,
            key="code_discount"
        )
    
    with col3:
        min_profit = st.number_input(
            "📊 الحد الأدنى للربح (%):",
            min_value=0.0,
            max_value=100.0,
            value=15.0,
            step=1.0,
            key="min_profit"
        )
    
    # المرحلة 3: إضافة المنتجات
    st.markdown("#### 3️⃣ إضافة المنتجات")
    
    # تحميل قائمة المنتجات من options.json
    products_list = load_products_list()
    
    if 'pricing_products' not in st.session_state:
        st.session_state.pricing_products = []
    
    # إضافة منتج جديد
    with st.expander("➕ إضافة منتج"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            product_name = st.selectbox(
                "المنتج:",
                products_list,
                key="new_product_name"
            )
        
        # جلب بيانات التسعير التلقائية
        pricing_data = get_product_pricing(product_name) if product_name else None
        default_base_price = pricing_data["base_price"] if pricing_data else 100.0
        
        with col2:
            base_price = st.number_input(
                "السعر الأساسي:",
                min_value=0.0,
                value=float(default_base_price),
                step=1.0,
                key="new_base_price",
                help="يتم ملؤه تلقائياً من قاعدة البيانات"
            )
        
        with col3:
            cost = st.number_input(
                "التكلفة:",
                min_value=0.0,
                value=50.0,
                step=1.0,
                key="new_cost"
            )
        
        if st.button("➕ إضافة المنتج", use_container_width=True, type="primary"):
            # حساب الأسعار
            after_discount = base_price * (1 - base_discount / 100)
            after_code = after_discount * (1 - code_discount / 100)
            discount_percent = ((base_price - after_code) / base_price) * 100
            net_profit = after_code - cost
            profit_margin = (net_profit / after_code) * 100 if after_code > 0 else 0
            
            # تحديد الحالة
            if profit_margin >= 30:
                status = "ممتاز"
            elif profit_margin >= 15:
                status = "جيد"
            else:
                status = "تحذير"
            
            product = {
                "name": product_name,
                "base_price": base_price,
                "after_discount": after_discount,
                "after_code": after_code,
                "cost": cost,
                "profit_margin": profit_margin,
                "discount_percent": discount_percent,
                "net_profit": net_profit,
                "status": status
            }
            
            st.session_state.pricing_products.append(product)
            st.success(f"✅ تم إضافة {product_name}")
            st.rerun()
    
    # عرض المنتجات المضافة مع إمكانية التعديل
    if st.session_state.pricing_products:
        st.markdown("#### 📦 المنتجات المضافة")
        st.markdown("💡 **يمكنك تعديل أي قيمة في الجدول أدناه وسيتم إعادة الحساب تلقائياً**")
        
        # عرض كل منتج مع إمكانية التعديل
        for idx, product in enumerate(st.session_state.pricing_products):
            with st.expander(f"📦 {product['name']}", expanded=True):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    new_base_price = st.number_input(
                        "السعر الأساسي:",
                        min_value=0.0,
                        value=float(product['base_price']),
                        step=1.0,
                        key=f"edit_base_price_{idx}"
                    )
                
                with col2:
                    new_cost = st.number_input(
                        "التكلفة:",
                        min_value=0.0,
                        value=float(product['cost']),
                        step=1.0,
                        key=f"edit_cost_{idx}"
                    )
                
                with col3:
                    if st.button("🔄 تحديث", key=f"update_{idx}", use_container_width=True):
                        # إعادة حساب القيم
                        after_discount = new_base_price * (1 - base_discount / 100)
                        after_code = after_discount * (1 - code_discount / 100)
                        discount_percent = ((new_base_price - after_code) / new_base_price) * 100 if new_base_price > 0 else 0
                        net_profit = after_code - new_cost
                        profit_margin = (net_profit / after_code) * 100 if after_code > 0 else 0
                        
                        # تحديد الحالة
                        if profit_margin >= 30:
                            status = "ممتاز"
                        elif profit_margin >= 15:
                            status = "جيد"
                        else:
                            status = "تحذير"
                        
                        # تحديث المنتج
                        st.session_state.pricing_products[idx] = {
                            "name": product['name'],
                            "base_price": new_base_price,
                            "after_discount": after_discount,
                            "after_code": after_code,
                            "cost": new_cost,
                            "profit_margin": profit_margin,
                            "discount_percent": discount_percent,
                            "net_profit": net_profit,
                            "status": status
                        }
                        
                        st.success("✅ تم تحديث المنتج!")
                        st.rerun()
                
                # عرض النتائج المحسوبة
                st.markdown("---")
                col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                
                with col_r1:
                    st.metric("بعد الخصم", f"{product['after_discount']:.2f} ر.س")
                
                with col_r2:
                    st.metric("بعد الكود", f"{product['after_code']:.2f} ر.س")
                
                with col_r3:
                    st.metric("الربح الصافي", f"{product['net_profit']:.2f} ر.س")
                
                with col_r4:
                    # تلوين نسبة الربح
                    if product['profit_margin'] >= 30:
                        color = "green"
                    elif product['profit_margin'] >= 15:
                        color = "orange"
                    else:
                        color = "red"
                    
                    st.metric("نسبة الربح", f"{product['profit_margin']:.2f}%")
                    st.markdown(f"<p style='color: {color}; font-weight: bold; text-align: center;'>{product['status']}</p>", unsafe_allow_html=True)
                
                # زر الحذف
                if st.button("🗑️ حذف المنتج", key=f"delete_product_{idx}", type="secondary"):
                    st.session_state.pricing_products.pop(idx)
                    st.success("✅ تم حذف المنتج!")
                    st.rerun()
        
        # جدول ملخص
        st.markdown("---")
        st.markdown("#### 📊 جدول ملخص")
        
        df = pd.DataFrame(st.session_state.pricing_products)
        
        # تنسيق الأعمدة
        df_display = df.copy()
        df_display['base_price'] = df_display['base_price'].apply(lambda x: f"{x:.2f}")
        df_display['after_discount'] = df_display['after_discount'].apply(lambda x: f"{x:.2f}")
        df_display['after_code'] = df_display['after_code'].apply(lambda x: f"{x:.2f}")
        df_display['cost'] = df_display['cost'].apply(lambda x: f"{x:.2f}")
        df_display['profit_margin'] = df_display['profit_margin'].apply(lambda x: f"{x:.2f}%")
        df_display['discount_percent'] = df_display['discount_percent'].apply(lambda x: f"{x:.2f}%")
        df_display['net_profit'] = df_display['net_profit'].apply(lambda x: f"{x:.2f}")
        
        # تغيير أسماء الأعمدة للعربية
        df_display = df_display.rename(columns={
            'name': 'المنتج',
            'base_price': 'السعر الأساسي',
            'after_discount': 'بعد الخصم',
            'after_code': 'بعد الكود',
            'cost': 'التكلفة',
            'profit_margin': 'نسبة الربح',
            'discount_percent': 'نسبة الخصم',
            'net_profit': 'الربح الصافي',
            'status': 'الحالة'
        })
        
        st.dataframe(df_display, use_container_width=True, height=300)
        
        # الإحصائيات
        st.markdown("#### 📊 الإحصائيات")
        
        total_products = len(st.session_state.pricing_products)
        total_revenue = sum([p['after_code'] for p in st.session_state.pricing_products])
        total_cost = sum([p['cost'] for p in st.session_state.pricing_products])
        total_profit = total_revenue - total_cost
        avg_profit_margin = sum([p['profit_margin'] for p in st.session_state.pricing_products]) / total_products if total_products > 0 else 0
        avg_discount = sum([p['discount_percent'] for p in st.session_state.pricing_products]) / total_products if total_products > 0 else 0
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("📦 إجمالي المنتجات", total_products)
            st.metric("💰 الإيرادات المتوقعة", f"{total_revenue:,.2f} ر.س")
        
        with col2:
            st.metric("💵 إجمالي التكلفة", f"{total_cost:,.2f} ر.س")
            st.metric("📈 الربح المتوقع", f"{total_profit:,.2f} ر.س")
        
        with col3:
            st.metric("📊 متوسط الربح", f"{avg_profit_margin:.2f}%")
            st.metric("🎯 متوسط الخصم", f"{avg_discount:.2f}%")
        
        # النصائح الذكية من GPT
        st.markdown("#### 🤖 النصائح الذكية")
        
        if st.button("✨ احصل على نصائح من الذكاء الاصطناعي", use_container_width=True, type="primary"):
            with st.spinner("🤖 جاري تحليل الأسعار..."):
                advice = get_ai_pricing_advice(st.session_state.pricing_products, base_discount, code_discount)
                st.markdown(advice)
        
        # حفظ الخطة
        st.markdown("---")
        
        col1, col2 = st.columns([3, 1])
        
        with col2:
            if st.button("💾 حفظ الخطة", use_container_width=True, type="primary"):
                if not plan_name:
                    st.error("❌ يرجى إدخال اسم الخطة")
                else:
                    # إنشاء الخطة
                    new_plan = {
                        "id": f"plan_{datetime.now().strftime('%Y%m%d%H%M%S')}",
                        "name": plan_name,
                        "description": plan_description,
                        "campaign": selected_campaign if selected_campaign != "بدون ربط بحملة" else None,
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "created_by": st.session_state.get('username', 'مجهول'),
                        "settings": {
                            "base_discount": base_discount,
                            "code_discount": code_discount,
                            "min_profit_margin": min_profit
                        },
                        "products": st.session_state.pricing_products,
                        "analytics": {
                            "total_products": total_products,
                            "total_revenue": total_revenue,
                            "total_cost": total_cost,
                            "total_profit": total_profit,
                            "avg_profit_margin": avg_profit_margin,
                            "avg_discount": avg_discount
                        }
                    }
                    
                    # حفظ الخطة
                    pricing_plans = load_pricing_plans()
                    pricing_plans["plans"].append(new_plan)
                    
                    if save_pricing_plans(pricing_plans):
                        st.success("✅ تم حفظ خطة التسعير بنجاح!")
                        st.balloons()
                        st.session_state.pricing_products = []
                        st.rerun()
    else:
        st.info("📭 لم تتم إضافة أي منتجات بعد. ابدأ بإضافة المنتجات أعلاه.")


def load_campaigns():
    """تحميل الحملات من الملف"""
    try:
        if os.path.exists("moraselaty_campaigns.json"):
            with open("moraselaty_campaigns.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                # تحويل البيانات إلى الصيغة المتوقعة
                if isinstance(data, list):
                    return {"campaigns": data}
                return data if "campaigns" in data else {"campaigns": []}
        return {"campaigns": []}
    except Exception as e:
        st.error(f"خطأ في تحميل الحملات: {str(e)}")
        return {"campaigns": []}


def load_products_list():
    """تحميل قائمة المنتجات من products_pricing.json"""
    try:
        if os.path.exists("products_pricing.json"):
            with open("products_pricing.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                products = data.get("products", [])
                # إرجاع قائمة بأسماء المنتجات فقط
                return [p["name"] for p in products if "name" in p]
        # إذا لم يوجد الملف، جرب options.json
        elif os.path.exists("options.json"):
            with open("options.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("products", [])
        return []
    except Exception as e:
        st.error(f"خطأ في تحميل المنتجات: {str(e)}")
        return []


def get_product_pricing(product_name):
    """الحصول على بيانات التسعير لمنتج معين"""
    try:
        if os.path.exists("products_pricing.json"):
            with open("products_pricing.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                products = data.get("products", [])
                for p in products:
                    if p.get("name") == product_name:
                        return {
                            "base_price": p.get("base_price", 0),
                            "after_discount": p.get("after_discount", 0),
                            "after_code": p.get("after_code", 0)
                        }
        return None
    except:
        return None


def get_ai_pricing_advice(products, base_discount, code_discount):
    """الحصول على نصائح ذكية من GPT حول التسعير"""
    
    try:
        client = OpenAI()
        
        # إعداد البيانات للتحليل
        products_summary = []
        for p in products:
            products_summary.append({
                "name": p['name'],
                "base_price": p['base_price'],
                "final_price": p['after_code'],
                "cost": p['cost'],
                "profit_margin": p['profit_margin']
            })
        
        prompt = f"""
أنت خبير في التسعير والتسويق. قم بتحليل خطة التسعير التالية وقدم نصائح ذكية:

**إعدادات الخصم:**
- نسبة الخصم الأساسية: {base_discount}%
- نسبة خصم الكود: {code_discount}%

**المنتجات:**
{json.dumps(products_summary, ensure_ascii=False, indent=2)}

قدم تحليلاً شاملاً يتضمن:
1. تحليل فروقات الأسعار بين الأحجام المختلفة للمنتج الواحد (إن وجد)
2. نصائح حول التسعير النفسي
3. تحليل هوامش الربح وتحديد المنتجات التي تحتاج تعديل
4. اقتراحات لتحسين الاستراتيجية العامة
5. اقتراح باقات مربحة (إن أمكن)

اجعل الرد باللغة العربية، منظماً، واحترافياً مع استخدام الرموز التعبيرية.
"""
        
        response = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[
                {"role": "system", "content": "أنت خبير في التسعير والتسويق، تقدم نصائح احترافية ومدروسة."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=2000
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        return f"❌ خطأ في الحصول على النصائح: {str(e)}"


def export_plan_to_excel(plan):
    """تصدير خطة التسعير إلى Excel مع دعم كامل للعربية"""
    
    # إنشاء workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "خطة التسعير"
    
    # تعيين اتجاه الصفحة من اليمين لليسار
    ws.sheet_view.rightToLeft = True
    
    # الأنماط
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    yellow_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # معلومات الخطة
    ws['A1'] = "خطة التسعير"
    ws['A1'].font = Font(name="Arial", size=16, bold=True)
    ws['A1'].alignment = right_alignment
    ws.merge_cells('A1:I1')
    
    ws['A2'] = f"الاسم: {plan['name']}"
    ws['A2'].alignment = right_alignment
    ws.merge_cells('A2:I2')
    
    ws['A3'] = f"التاريخ: {plan.get('created_at', 'غير محدد')}"
    ws['A3'].alignment = right_alignment
    ws.merge_cells('A3:I3')
    
    # الإحصائيات
    ws['A5'] = "📊 الإحصائيات"
    ws['A5'].font = Font(name="Arial", size=14, bold=True)
    ws['A5'].alignment = right_alignment
    ws.merge_cells('A5:I5')
    
    stats = [
        ["إجمالي المنتجات", plan['analytics']['total_products']],
        ["الإيرادات المتوقعة", f"{plan['analytics']['total_revenue']:,.2f} ر.س"],
        ["إجمالي التكلفة", f"{plan['analytics']['total_cost']:,.2f} ر.س"],
        ["الربح المتوقع", f"{plan['analytics']['total_profit']:,.2f} ر.س"],
        ["متوسط الربح", f"{plan['analytics']['avg_profit_margin']:.2f}%"],
        ["متوسط الخصم", f"{plan['analytics']['avg_discount']:.2f}%"]
    ]
    
    row = 6
    for stat in stats:
        ws[f'A{row}'] = stat[0]
        ws[f'B{row}'] = stat[1]
        ws[f'A{row}'].alignment = right_alignment
        ws[f'B{row}'].alignment = center_alignment
        row += 1
    
    # جدول المنتجات
    ws[f'A{row+1}'] = "📦 المنتجات"
    ws[f'A{row+1}'].font = Font(name="Arial", size=14, bold=True)
    ws[f'A{row+1}'].alignment = right_alignment
    ws.merge_cells(f'A{row+1}:I{row+1}')
    
    # رؤوس الأعمدة
    row += 2
    headers = ['المنتج', 'السعر الأساسي', 'بعد الخصم', 'بعد الكود', 'التكلفة', 'نسبة الربح', 'نسبة الخصم', 'الربح الصافي', 'الحالة']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # بيانات المنتجات
    for product in plan['products']:
        row += 1
        
        # تحديد لون الخلفية بناءً على نسبة الربح
        if product['profit_margin'] >= 30:
            fill = green_fill
        elif product['profit_margin'] >= 15:
            fill = yellow_fill
        else:
            fill = red_fill
        
        data = [
            product['name'],
            f"{product['base_price']:.2f}",
            f"{product['after_discount']:.2f}",
            f"{product['after_code']:.2f}",
            f"{product['cost']:.2f}",
            f"{product['profit_margin']:.2f}%",
            f"{product['discount_percent']:.2f}%",
            f"{product['net_profit']:.2f}",
            product['status']
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.alignment = center_alignment
            cell.border = border
            
            # تلوين خلية نسبة الربح
            if col_idx == 6:
                cell.fill = fill
    
    # ضبط عرض الأعمدة
    column_widths = [25, 15, 15, 15, 15, 15, 15, 15, 12]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # حفظ الملف في الذاكرة
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

